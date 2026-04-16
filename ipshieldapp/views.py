# ===============================================
# IMPORTS
# ===============================================
from django.core.mail import send_mail
from django.conf import settings
from decimal import Decimal
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q
from django.http import FileResponse, Http404
from django.db import IntegrityError
import os
import openpyxl
from django.http import HttpResponse
from .models import CustomerActivityLog
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from .decorators import customer_login_required

from .models import *
from .forms import *
from django.utils import timezone

# ===============================================
# EMAIL HELPER
# ===============================================
def _send_email(to_email, subject, message):
    if not to_email:
        return
    try:
        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[to_email],
            fail_silently=False,
        )
        print(f"✅ Email sent → {to_email}")
    except Exception as e:
        print(f"❌ Email error: {e}")

def _log_activity(customer, action, note='', request=None):
    ip = None
    if request:
        x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
        ip = x_forwarded.split(',')[0] if x_forwarded else request.META.get('REMOTE_ADDR')
    CustomerActivityLog.objects.create(
        customer=customer, action=action, note=note, ip_address=ip
    )
def lock_contract_fields(contract_form):
    for field_name, field in contract_form.fields.items():
        field.disabled = True
        field.widget.attrs['readonly'] = True
        field.widget.attrs['class'] = field.widget.attrs.get('class', '') + ' readonly-field'


# ===============================================
# CUSTOMER LIST + SEARCH
# ===============================================
def home(request):
    import datetime
    from django.db.models import Count
    q = request.GET.get('q', '').strip()
    today = datetime.date.today()
    year_range  = range(today.year - 3, today.year + 1)
    month_range = range(1, 13)

    # ── Xuất theo ngày ──
    if request.GET.get('export') == 'excel_by_date':
        export_date = request.GET.get('export_date', str(today))
        export_logs = CustomerActivityLog.objects.select_related('customer').filter(
            created_at__date=export_date
        ).order_by('customer__customer_code', 'created_at')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hoạt động theo ngày"
        ws.append(['STT', 'Mã KH', 'Tên KH', 'Hành động', 'Ghi chú', 'IP', 'Thời gian'])
        for idx, log in enumerate(export_logs, 1):
            ws.append([idx, log.customer.customer_code, log.customer.name,
                       log.get_action_display(), log.note or '',
                       log.ip_address or '', log.created_at.strftime('%H:%M:%S — %d/%m/%Y')])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="hoat_dong_ngay_{export_date}.xlsx"'
        wb.save(response)
        return response

    # ── Xuất theo KH ──
    if request.GET.get('export') == 'excel_by_customer':
        export_date = request.GET.get('export_date', str(today))
        export_customer_id = request.GET.get('export_customer_id', '').strip()
        export_logs = CustomerActivityLog.objects.select_related('customer').filter(
            created_at__date=export_date
        ).order_by('created_at')
        if export_customer_id:
            export_logs = export_logs.filter(customer__id=export_customer_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Theo KH"
        ws.append(['STT', 'Mã KH', 'Tên KH', 'Hành động', 'Ghi chú', 'IP', 'Thời gian'])
        for idx, log in enumerate(export_logs, 1):
            ws.append([idx, log.customer.customer_code, log.customer.name,
                       log.get_action_display(), log.note or '',
                       log.ip_address or '', log.created_at.strftime('%H:%M:%S — %d/%m/%Y')])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="hoat_dong_kh_{export_customer_id}_{export_date}.xlsx"'
        wb.save(response)
        return response

    customers_all = Customer.objects.all()
    if q:
        customers_all = customers_all.filter(
            Q(customer_code__icontains=q) | Q(name__icontains=q) |
            Q(email__icontains=q) | Q(phone__icontains=q)
        )

    # Danh sách KH cho modal export (chỉ KH có activity log)
    export_customers = CustomerActivityLog.objects.select_related('customer') \
        .values('customer__id', 'customer__customer_code', 'customer__name') \
        .distinct().order_by('customer__customer_code')

    sliders = Slider.objects.filter(is_active=True)
    mascots = Mascot.objects.filter(is_active=True)
    nhanhieudocquyen = NhanHieuDocQuyen.objects.filter(is_active=True).order_by('id')

    return render(request, 'khachhang.html', {
        'customers': customers_all,
        'export_customers': export_customers,   # <-- cho modal
        'sliders': sliders,
        'mascots': mascots,
        'nhanhieudocquyen': nhanhieudocquyen,
        'year_range': year_range,
        'month_range': month_range,
        'current_month': today.month,
        'current_year':  today.year,
        'today': str(today),
        'q': q,
    })


# ===============================================
# ADD CONTRACT
# ===============================================
def add_contract(request):
    if request.method == 'POST':
        print("\n" + "=" * 60)
        print("🔍 POST REQUEST RECEIVED")
        print("=" * 60)

        print("\n📋 POST Data Keys:")
        for key in request.POST.keys():
            if 'trademark' in key or 'copyright' in key:
                print(f"  {key}: {request.POST.get(key)}")

        contract_form = ContractForm(request.POST, request.FILES)

        if not contract_form.is_valid():
            print("\n❌ Contract form invalid:")
            print(contract_form.errors)
            for field, errors in contract_form.errors.items():
                for error in errors:
                    field_label = contract_form.fields.get(field).label if field in contract_form.fields else field
                    messages.error(request, f"{field_label}: {error}")

            return render(request, "add_contract.html", {
                'contract_form': contract_form,
                'trademark_formset': TrademarkFormSet(request.POST, request.FILES, prefix='trademark'),
                'copyright_formset': CopyrightFormSet(request.POST, request.FILES, prefix='copyright'),
                'business_form': BusinessRegistrationForm(request.POST, request.FILES),
                'investment_form': InvestmentForm(request.POST, request.FILES),
                'other_form': OtherServiceForm(request.POST, request.FILES),
            })

        print("✅ Contract form valid")

        contract = contract_form.save(commit=False)
        contract.status = 'completed' if contract.payment_type == 'full' else 'processing'

        try:
            contract.save()
            print(f"✅ Contract saved: {contract.contract_no}")

            # 📷 LƯU ẢNH CHỤP HỢP ĐỒNG
            contract_images = request.FILES.getlist('contract_images')
            for f in contract_images:
                ContractImage.objects.create(contract=contract, image=f, name=f.name)
            service_type = contract.service_type
            print(f"\n📦 Processing service type: {service_type}")

            if service_type == 'nhanhieu':
                print("\n🏷️ Processing TRADEMARK formset...")
                trademark_formset = TrademarkFormSet(request.POST, request.FILES, prefix='trademark')

                if not trademark_formset.is_valid():
                    contract.delete()
                    for idx, form_errors in enumerate(trademark_formset.errors):
                        if form_errors:
                            for field, errors in form_errors.items():
                                for error in errors:
                                    messages.error(request, f"Nhãn hiệu #{idx + 1} - {field}: {error}")
                    for error in trademark_formset.non_form_errors():
                        messages.error(request, f"Lỗi formset: {error}")
                    return render(request, "add_contract.html", {
                        'contract_form': contract_form,
                        'trademark_formset': trademark_formset,
                        'copyright_formset': CopyrightFormSet(prefix='copyright', queryset=CopyrightService.objects.none()),
                        'business_form': BusinessRegistrationForm(),
                        'investment_form': InvestmentForm(),
                        'other_form': OtherServiceForm(),
                    })

                valid_forms = [
                    form for form in trademark_formset
                    if form.cleaned_data and not form.cleaned_data.get('DELETE', False)
                ]

                if len(valid_forms) == 0:
                    messages.warning(request, "⚠️ Hợp đồng đã lưu nhưng chưa có thông tin nhãn hiệu")
                else:
                    saved_count = 0
                    for idx, form in enumerate(valid_forms):
                        instance = form.save(commit=False)
                        instance.contract = contract
                        instance.save()
                        saved_count += 1
                        files = request.FILES.getlist(f'trademark_files_{idx}')
                        for f in files:
                            Certificate.objects.create(content_object=instance, file=f, name=f.name)
                    print(f"✅ Saved {saved_count} trademarks")

            elif service_type == 'banquyen':
                print("\n©️ Processing COPYRIGHT formset...")
                copyright_formset = CopyrightFormSet(request.POST, request.FILES, prefix='copyright')

                if not copyright_formset.is_valid():
                    contract.delete()
                    for idx, form_errors in enumerate(copyright_formset.errors):
                        if form_errors:
                            for field, errors in form_errors.items():
                                for error in errors:
                                    messages.error(request, f"Bản quyền #{idx + 1} - {field}: {error}")
                    for error in copyright_formset.non_form_errors():
                        messages.error(request, f"Lỗi formset: {error}")
                    return render(request, "add_contract.html", {
                        'contract_form': contract_form,
                        'trademark_formset': TrademarkFormSet(prefix='trademark', queryset=TrademarkService.objects.none()),
                        'copyright_formset': copyright_formset,
                        'business_form': BusinessRegistrationForm(),
                        'investment_form': InvestmentForm(),
                        'other_form': OtherServiceForm(),
                    })

                valid_forms = [
                    form for form in copyright_formset
                    if form.cleaned_data and not form.cleaned_data.get('DELETE', False)
                ]

                if len(valid_forms) == 0:
                    messages.warning(request, "⚠️ Hợp đồng đã lưu nhưng chưa có thông tin bản quyền")
                else:
                    saved_count = 0
                    for idx, form in enumerate(valid_forms):
                        instance = form.save(commit=False)
                        instance.contract = contract
                        instance.save()
                        saved_count += 1
                        files = request.FILES.getlist(f'copyright_files_{idx}')
                        for f in files:
                            Certificate.objects.create(content_object=instance, file=f, name=f.name)
                    print(f"✅ Saved {saved_count} copyrights")

            elif service_type == 'dkkd':
                form = BusinessRegistrationForm(request.POST, request.FILES)
                if not form.is_valid():
                    contract.delete()
                    for field, errors in form.errors.items():
                        for error in errors:
                            field_label = form.fields.get(field).label if field in form.fields else field
                            messages.error(request, f"ĐKKD - {field_label}: {error}")
                    return render(request, "add_contract.html", {
                        'contract_form': contract_form,
                        'trademark_formset': TrademarkFormSet(prefix='trademark', queryset=TrademarkService.objects.none()),
                        'copyright_formset': CopyrightFormSet(prefix='copyright', queryset=CopyrightService.objects.none()),
                        'business_form': form,
                        'investment_form': InvestmentForm(),
                        'other_form': OtherServiceForm(),
                    })
                if any(form.cleaned_data.values()):
                    obj = form.save(commit=False)
                    obj.contract = contract
                    obj.save()
                    files = request.FILES.getlist('business_files')
                    for f in files:
                        Certificate.objects.create(content_object=obj, file=f, name=f.name)
                else:
                    messages.warning(request, "⚠️ Hợp đồng đã lưu nhưng chưa có thông tin ĐKKD")

            elif service_type == 'dautu':
                form = InvestmentForm(request.POST, request.FILES)
                if not form.is_valid():
                    contract.delete()
                    for field, errors in form.errors.items():
                        for error in errors:
                            field_label = form.fields.get(field).label if field in form.fields else field
                            messages.error(request, f"Đầu tư - {field_label}: {error}")
                    return render(request, "add_contract.html", {
                        'contract_form': contract_form,
                        'trademark_formset': TrademarkFormSet(prefix='trademark', queryset=TrademarkService.objects.none()),
                        'copyright_formset': CopyrightFormSet(prefix='copyright', queryset=CopyrightService.objects.none()),
                        'business_form': BusinessRegistrationForm(),
                        'investment_form': form,
                        'other_form': OtherServiceForm(),
                    })
                if any(form.cleaned_data.values()):
                    obj = form.save(commit=False)
                    obj.contract = contract
                    obj.save()
                    files = request.FILES.getlist('investment_files')
                    for f in files:
                        Certificate.objects.create(content_object=obj, file=f, name=f.name)
                else:
                    messages.warning(request, "⚠️ Hợp đồng đã lưu nhưng chưa có thông tin đầu tư")

            else:  # khac
                form = OtherServiceForm(request.POST, request.FILES)
                if not form.is_valid():
                    contract.delete()
                    for field, errors in form.errors.items():
                        for error in errors:
                            field_label = form.fields.get(field).label if field in form.fields else field
                            messages.error(request, f"Dịch vụ khác - {field_label}: {error}")
                    return render(request, "add_contract.html", {
                        'contract_form': contract_form,
                        'trademark_formset': TrademarkFormSet(prefix='trademark', queryset=TrademarkService.objects.none()),
                        'copyright_formset': CopyrightFormSet(prefix='copyright', queryset=CopyrightService.objects.none()),
                        'business_form': BusinessRegistrationForm(),
                        'investment_form': InvestmentForm(),
                        'other_form': form,
                    })
                if any(form.cleaned_data.values()):
                    obj = form.save(commit=False)
                    obj.contract = contract
                    obj.save()
                    files = request.FILES.getlist('other_files')
                    for f in files:
                        Certificate.objects.create(content_object=obj, file=f, name=f.name)
                else:
                    messages.warning(request, "⚠️ Hợp đồng đã lưu nhưng chưa có thông tin dịch vụ")

            if contract.payment_type == 'installment':
                contract.create_installments()
            else:
                contract.status = 'completed'
                contract.save(update_fields=['status'])

            customer = contract.customer
            customer.status = 'pending'
            customer.save()

            messages.success(request, "✅ Tạo hợp đồng thành công!")

            # 🔔 GỬI MAIL HỢP ĐỒNG MỚI
            _send_email(
                to_email=contract.customer.email,
                subject=f"[IPSHIELD] Hợp đồng mới – {contract.contract_no}",
                message=f"""Kính gửi {contract.customer.name},

Hợp đồng của bạn đã được tạo thành công:

  • Số hợp đồng : {contract.contract_no}
  • Dịch vụ     : {contract.get_service_type_display()}
  • Giá trị     : {contract.contract_value:,.0f} VNĐ
  • Hình thức TT: {contract.get_payment_type_display()}

Vui lòng đăng nhập portal để xem chi tiết.

Trân trọng,
IPShield
""",
            )

            return redirect('contract_detail', id=contract.id)

        except Exception as e:
            import traceback
            print(f"Error: {str(e)}")
            print(traceback.format_exc())
            if contract.id:
                contract.delete()
            messages.error(request, f"❌ Có lỗi xảy ra: {str(e)}")
            return redirect('add_contract')

    print("\n📄 GET REQUEST - Rendering empty form")
    return render(request, "add_contract.html", {
        'contract_form': ContractForm(),
        'trademark_formset': TrademarkFormSet(prefix='trademark', queryset=TrademarkService.objects.none()),
        'copyright_formset': CopyrightFormSet(prefix='copyright', queryset=CopyrightService.objects.none()),
        'business_form': BusinessRegistrationForm(),
        'investment_form': InvestmentForm(),
        'other_form': OtherServiceForm(),
    })


def contract_delete(request, pk):
    print("🔥 DELETE VIEW CALLED:", pk)
    contract = get_object_or_404(Contract, pk=pk)
    contract.delete()
    return redirect('contract_list')

def upload_contract_image(request):
    if request.method == 'POST':
        contract_id = request.POST.get('contract_id')
        contract = get_object_or_404(Contract, id=contract_id)
        name = request.POST.get('name', '').strip()
        files = request.FILES.getlist('images')
        if not files:
            messages.error(request, 'Chưa chọn ảnh')
            return redirect(request.META.get('HTTP_REFERER'))
        for f in files:
            ContractImage.objects.create(contract=contract, image=f, name=name or f.name)
        messages.success(request, f'Đã tải lên {len(files)} ảnh hợp đồng')
        return redirect(request.META.get('HTTP_REFERER'))


def delete_contract_image(request, pk):
    img = get_object_or_404(ContractImage, pk=pk)
    contract_id = img.contract.id
    if img.image:
        img.image.delete(save=False)
    img.delete()
    messages.success(request, '✅ Đã xóa ảnh')
    return redirect('contract_detail', id=contract_id)
# ===============================================
# CUSTOMER DETAIL
# ===============================================
def customer_detail(request, id):
    customer = get_object_or_404(Customer, id=id)
    contracts = customer.contracts.all()
    return render(request, 'customer_detail.html', {'customer': customer, 'contracts': contracts})


def customer_change_status(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerStatusForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Cập nhật trạng thái thành công!")
            return redirect('customer_detail', id=customer.id)
    else:
        form = CustomerStatusForm(instance=customer)
    return render(request, 'customer_change_status.html', {'customer': customer, 'form': form})


def customer_edit(request, id):
    customer = get_object_or_404(Customer, id=id)
    if request.method == 'POST':
        form = CustomerForm(request.POST, request.FILES, instance=customer)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "✅ Cập nhật khách hàng thành công!")
                return redirect('customer_detail', id=id)
            except IntegrityError:
                customer_code = form.cleaned_data.get('customer_code', '')
                messages.error(request, f"⚠️ Mã khách hàng '{customer_code}' đã tồn tại!")
            except Exception as e:
                messages.error(request, f"❌ Có lỗi xảy ra: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields.get(field).label if field in form.fields else field}: {error}")
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'customer_edit.html', {'form': form, 'customer': customer})


def customer_delete(request, id):
    customer = get_object_or_404(Customer, id=id)
    if request.method == "POST":
        customer.delete()
        messages.success(request, "✅ Đã xóa khách hàng!")
        return redirect('home')
    return render(request, 'customer_delete_confirm.html', {'customer': customer})


# ===============================================
# CONTRACT LIST
# ===============================================
def contract_list(request):
    contracts = Contract.objects.select_related("customer").order_by("-created_at")
    return render(request, "contract_list.html", {"contracts": contracts})


# ===============================================
# CONTRACT DETAIL
# ===============================================
from datetime import date, datetime
from django.db.models import QuerySet


def contract_detail(request, id):
    contract = get_object_or_404(Contract, id=id)
    installments = contract.installments.all()
    paid_count = installments.filter(is_paid=True).count()

    service = None
    if contract.service_type == "nhanhieu":
        service = contract.trademarks.all()
    elif contract.service_type == "banquyen":
        service = contract.copyrights.all()
    elif contract.service_type == "dkkd":
        service = BusinessRegistrationService.objects.filter(contract=contract)
    elif contract.service_type == "dautu":
        service = InvestmentService.objects.filter(contract=contract)
    else:
        service = OtherService.objects.filter(contract=contract)

    if contract.payment_type == "full" and contract.status != "completed":
        contract.status = "completed"
        contract.save(update_fields=["status"])

    if request.method == "POST":
        action = request.POST.get("action")

        # 1️⃣ ĐÁNH DẤU XUẤT HÓA ĐƠN
        if action == "log_toggle_bill":
            log_id = request.POST.get("log_id")
            log_entry = get_object_or_404(PaymentLog, id=log_id)
            if not log_entry.is_exported_bill:
                log_entry.is_exported_bill = True
                log_entry.bill_exported_at = timezone.now()
                log_entry.save()
                messages.success(request, "✅ Đã đánh dấu xuất hóa đơn")
            return redirect("contract_detail", id=contract.id)

        # 2️⃣ THANH TOÁN ĐỢT
        elif action == "pay_installment":
            installment_id = request.POST.get("installment_id")
            try:
                ins = contract.installments.get(id=installment_id)
                if ins.is_paid:
                    messages.warning(request, "⚠️ Đợt này đã được thanh toán rồi")
                    return redirect("contract_detail", id=contract.id)

                remaining = ins.amount - ins.paid_amount
                ins.paid_amount = ins.amount
                ins.is_paid = True
                ins.paid_date = timezone.now().date()
                ins.save()

                PaymentLog.objects.create(
                    contract=contract,
                    installment=ins,
                    amount_paid=remaining,
                    paid_at=timezone.now()
                )

                if contract.remaining_amount <= 0:
                    contract.status = "completed"
                    contract.save(update_fields=["status"])

                messages.success(request, f"✅ Đã thanh toán đợt {ins.notes} - Số tiền: {remaining:,.0f} VNĐ")

                # 🔔 GỬI MAIL THANH TOÁN ĐỢT
                _send_email(
                    to_email=contract.customer.email,
                    subject=f"[IPSHIELD] Xác nhận thanh toán – {contract.contract_no}",
                    message=f"""Kính gửi {contract.customer.name},

Đợt thanh toán "{ins.notes}" của hợp đồng {contract.contract_no} đã được ghi nhận.

  • Số tiền : {remaining:,.0f} VNĐ
  • Còn lại : {contract.remaining_amount:,.0f} VNĐ

Trân trọng,
IPShield
""",
                )

                return redirect("contract_detail", id=contract.id)

            except PaymentInstallment.DoesNotExist:
                messages.error(request, "❌ Không tìm thấy đợt thanh toán")
                return redirect("contract_detail", id=contract.id)
            except Exception as e:
                messages.error(request, f"❌ Có lỗi xảy ra: {str(e)}")
                return redirect("contract_detail", id=contract.id)

        # 3️⃣ THANH TOÁN TỪNG PHẦN
        elif action == "partial_payment":
            installment_id = request.POST.get("installment_id")
            paid_amount_raw = request.POST.get("paid_amount")

            if not paid_amount_raw:
                messages.error(request, "❌ Vui lòng nhập số tiền thanh toán")
                return redirect("contract_detail", id=contract.id)

            try:
                paid_amount = Decimal(paid_amount_raw)
                ins = contract.installments.get(id=installment_id)

                ins.paid_amount += paid_amount
                if ins.paid_amount >= ins.amount:
                    ins.is_paid = True
                    ins.paid_date = timezone.now().date()
                ins.save()

                PaymentLog.objects.create(
                    contract=contract,
                    installment=ins,
                    amount_paid=paid_amount,
                    paid_at=timezone.now()
                )

                if contract.remaining_amount <= 0:
                    contract.status = "completed"
                    contract.save(update_fields=["status"])

                messages.success(request, f"✅ Đã ghi nhận thanh toán: {paid_amount:,.0f} VNĐ")

                # 🔔 GỬI MAIL THANH TOÁN TỪNG PHẦN
                _send_email(
                    to_email=contract.customer.email,
                    subject=f"[IPSHIELD] Xác nhận thanh toán – {contract.contract_no}",
                    message=f"""Kính gửi {contract.customer.name},

Thanh toán của hợp đồng {contract.contract_no} đã được ghi nhận.

  • Số tiền : {paid_amount:,.0f} VNĐ
  • Còn lại : {contract.remaining_amount:,.0f} VNĐ

Trân trọng,
IPShield
""",
                )

                return redirect("contract_detail", id=contract.id)

            except Exception as e:
                messages.error(request, f"❌ Có lỗi xảy ra: {str(e)}")
                return redirect("contract_detail", id=contract.id)

        # 4️⃣ XUẤT HÓA ĐƠN
        if action == 'export_bill':
            installment_id = request.POST.get('installment_id')
            try:
                installment = PaymentInstallment.objects.get(id=installment_id)
                if installment.is_paid:
                    installment.is_exported_bill = True
                    installment.bill_exported_at = timezone.now()
                    installment.save()
                    messages.success(request, '✅ Đã xuất hóa đơn đỏ!')
                else:
                    messages.error(request, '⚠️ Đợt này chưa thanh toán!')
            except PaymentInstallment.DoesNotExist:
                messages.error(request, '❌ Không tìm thấy đợt thanh toán!')
            return redirect('contract_detail', id=id)

    return render(request, "contract_detail.html", {
        "contract": contract,
        "service": service,
        "installments": installments,
        "paid_count": paid_count,
    })


from .forms import InstallmentFormSet


@login_required
def edit_installment_amounts(request, contract_id):
    contract = get_object_or_404(Contract, id=contract_id)

    if contract.payment_type != 'installment':
        messages.error(request, "❌ Hợp đồng này không phải trả góp")
        return redirect('contract_detail', id=contract_id)

    installments = contract.installments.all().order_by('due_date')

    if request.method == 'POST':
        formset = InstallmentFormSet(request.POST, queryset=installments)
        if formset.is_valid():
            total_entered = 0
            for form in formset:
                if form.cleaned_data:
                    amount = form.cleaned_data.get('amount', 0)
                    if amount:
                        total_entered += amount

            if total_entered != contract.contract_value:
                messages.warning(
                    request,
                    f"⚠️ Tổng số tiền các đợt ({total_entered:,.0f} VNĐ) "
                    f"khác với giá trị hợp đồng ({contract.contract_value:,.0f} VNĐ)"
                )

            instances = formset.save(commit=False)
            for instance in instances:
                instance.save()

            messages.success(request, "✅ Đã cập nhật số tiền các đợt thanh toán!")
            return redirect('contract_detail', id=contract_id)
        else:
            messages.error(request, "❌ Có lỗi trong form, vui lòng kiểm tra lại")
            print("Formset errors:", formset.errors)
            print("Non-form errors:", formset.non_form_errors())
    else:
        formset = InstallmentFormSet(queryset=installments)

    current_total = sum(ins.amount for ins in installments)
    return render(request, 'edit_installment_amounts.html', {
        'contract': contract,
        'formset': formset,
        'installments': installments,
        'current_total': current_total,
    })


# ===============================================
# CONTRACT EDIT
# ===============================================
@login_required
def contract_edit(request, id):
    contract = get_object_or_404(Contract, id=id)

    contract_form = None
    service_form = None
    service_formset = None
    FormSetClass = None
    ServiceForm = None
    queryset = None
    prefix = None
    service = None

    if contract.service_type == "nhanhieu":
        FormSetClass = TrademarkFormSet
        queryset = TrademarkService.objects.filter(contract=contract)
        prefix = "trademark"
    elif contract.service_type == "banquyen":
        FormSetClass = CopyrightFormSet
        queryset = CopyrightService.objects.filter(contract=contract)
        prefix = "copyright"
    elif contract.service_type == "dkkd":
        ServiceForm = BusinessRegistrationForm
        service = BusinessRegistrationService.objects.filter(contract=contract).first()
    elif contract.service_type == "dautu":
        ServiceForm = InvestmentForm
        service = InvestmentService.objects.filter(contract=contract).first()
    else:
        ServiceForm = OtherServiceForm
        service = OtherService.objects.filter(contract=contract).first()

    if request.method == "POST":
        contract_form = ContractForm(request.POST, instance=contract)
        lock_contract_fields(contract_form)

        # 🔔 DÙNG CHUNG CHO CẢ 2 NHÁNH
        def send_service_updated_email():
            _send_email(
                to_email=contract.customer.email,
                subject=f"[IPSHIELD] Hợp đồng {contract.contract_no} vừa được cập nhật",
                message=f"""Kính gửi {contract.customer.name},

Thông tin hợp đồng {contract.contract_no} vừa được chỉnh sửa.

Vui lòng đăng nhập portal để xem chi tiết.

Trân trọng,
IPShield
""",
            )

        # ===== FORMSET (NHÃN HIỆU / BẢN QUYỀN) =====
        if FormSetClass:
            service_formset = FormSetClass(
                request.POST, request.FILES, queryset=queryset, prefix=prefix
            )
            if contract_form.is_valid() and service_formset.is_valid():
                contract_form.save()
                instances = service_formset.save(commit=False)
                for obj in instances:
                    obj.contract = contract
                    obj.save()
                for obj in service_formset.deleted_objects:
                    obj.delete()

                send_service_updated_email()  # 🔔

                messages.success(request, "✅ Cập nhật hợp đồng thành công!")
                return redirect("contract_detail", id=contract.id)

        # ===== SERVICE ĐƠN (DKKD / ĐẦU TƯ / KHÁC) =====
        else:
            service_form = ServiceForm(request.POST, request.FILES, instance=service)
            if contract_form.is_valid() and service_form.is_valid():
                contract_form.save()
                obj = service_form.save(commit=False)
                obj.contract = contract
                obj.save()

                send_service_updated_email()  # 🔔

                messages.success(request, "✅ Cập nhật hợp đồng thành công!")
                return redirect("contract_detail", id=contract.id)

    else:
        contract_form = ContractForm(instance=contract)
        lock_contract_fields(contract_form)
        if FormSetClass:
            service_formset = FormSetClass(queryset=queryset, prefix=prefix)
        else:
            service_form = ServiceForm(instance=service)

    return render(request, "contract_edit.html", {
        "contract": contract,
        "contract_form": contract_form,
        "service_formset": service_formset,
        "service_form": service_form,
    })


# ===============================================
# DOWNLOAD CERTIFICATE
# ===============================================
def download_certificate(request, id):
    contract = get_object_or_404(Contract, id=id)
    if contract.service_type == "nhanhieu":
        service = contract.trademarks.first()
        if not service or not service.certificate_file:
            raise Http404("Không có giấy chứng nhận")
    elif contract.service_type == "banquyen":
        service = CopyrightService.objects.get(contract=contract)
    elif contract.service_type == "dkkd":
        service = BusinessRegistrationService.objects.get(contract=contract)
    elif contract.service_type == "dautu":
        service = InvestmentService.objects.get(contract=contract)
    else:
        service = OtherService.objects.get(contract=contract)

    if not service.certificate_file:
        raise Http404("Không có giấy chứng nhận")

    return FileResponse(
        service.certificate_file.open("rb"),
        as_attachment=True,
        filename=os.path.basename(service.certificate_file.name)
    )


def register_certificate(request, business_id):
    business = get_object_or_404(BusinessRegistrationService, id=business_id)
    file_field = business.registration_certificate
    if not file_field:
        raise Http404("Chưa có file đăng ký kinh doanh")
    file_path = file_field.path
    if not os.path.exists(file_path):
        raise Http404("File không tồn tại")
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=os.path.basename(file_path))


# ===============================================
# CUSTOMER CREATE
# ===============================================
def add_customer(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "✅ Thêm khách hàng thành công!")
                return redirect('home')
            except IntegrityError:
                messages.error(request, "⚠️ Mã khách hàng đã tồn tại!")
        else:
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, error)
    else:
        form = CustomerForm()
    return render(request, 'add_customer.html', {'form': form})


from django.http import JsonResponse


def search_customer(request):
    query = request.GET.get('q', '').strip()
    if len(query) < 1:
        return JsonResponse([], safe=False)
    try:
        customers = Customer.objects.filter(
            Q(customer_code__icontains=query) | Q(name__icontains=query)
        )[:10]
        results = [{
            'id': c.id,
            'code': c.customer_code,
            'name': c.name,
            'phone': str(c.phone) if c.phone else '',
            'email': c.email if c.email else ''
        } for c in customers]
        return JsonResponse(results, safe=False)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)


# ===============================================
# LOGIN / LOGOUT
# ===============================================
def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    if request.customer:
        return redirect('portal_dashboard')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()

        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('home')

        try:
            customer = Customer.objects.get(customer_code=username)
            if customer.check_password(password):
                request.session['customer_id'] = customer.id
                customer.last_login = timezone.now()
                customer.save(update_fields=['last_login'])
                return redirect('portal_dashboard')
            else:
                messages.error(request, '❌ Sai mật khẩu.')
        except Customer.DoesNotExist:
            messages.error(request, '❌ Tài khoản không tồn tại.')

    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ===============================================
# PROTECT VIEWS
# ===============================================
def protect_views(*views):
    for view in views:
        globals()[view.__name__] = login_required(view)


protect_views(
    home, add_contract, customer_detail, customer_change_status,
    customer_edit, customer_delete, contract_list, contract_detail,
    contract_edit, download_certificate, add_customer, search_customer,
)


# ===============================================
# CONTRACT SEARCH
# ===============================================
from django.shortcuts import render, redirect
from .models import Contract, TrademarkService


def contract_search(request):
    import datetime
    q = request.GET.get('q', '').strip()
    customers = Customer.objects.all().order_by('created_at')
    contracts = Contract.objects.filter(service_type='nhanhieu').order_by('-created_at')
    filled_count = contracts.filter(trademarks__filing_date__isnull=False).distinct().count()
    decision_count = contracts.filter(trademarks__decision_date__isnull=False).distinct().count()
    valid_count = contracts.filter(trademarks__valid_date__isnull=False).distinct().count()
    trademarks = []
    today = datetime.date.today()
    year_range  = range(today.year - 3, today.year + 1)
    month_range = range(1, 13)
    

    if q:
        trademarks = TrademarkService.objects.filter(
            app_no__icontains=q
        ).select_related('contract', 'contract__customer')
        if trademarks.exists():
            contract_ids = trademarks.values_list('contract_id', flat=True)
            contracts = Contract.objects.filter(id__in=contract_ids).order_by('-created_at')
        else:
            contracts = contracts.filter(
                Q(contract_no__icontains=q) |
                Q(customer__customer_code__icontains=q) |
                Q(customer__name__icontains=q)
            ).distinct()

    return render(request, 'contract_search.html', {'contracts': contracts, 'trademarks': trademarks, 'q': q,
                                                    'year_range': year_range,
                                                    'filled_count': filled_count,
                                                    'decision_count': decision_count,
                                                    'valid_count': valid_count,
                                                    'customers': customers,
        'month_range': month_range,
        'current_month':  today.month,
        'current_year':   today.year,
        'today': str(today),})


def contract_copyright_search(request):
    q = request.GET.get('q', '').strip()
    contracts = Contract.objects.filter(service_type='banquyen').order_by('-created_at')
    copyrights = []

    if q:
        copyrights = CopyrightService.objects.filter(
            certificate_no__icontains=q
        ).select_related('contract', 'contract__customer')
        if copyrights.exists():
            contract_ids = copyrights.values_list('contract_id', flat=True)
            contracts = Contract.objects.filter(id__in=contract_ids).order_by('-created_at')
        else:
            contracts = contracts.filter(
                Q(contract_no__icontains=q) |
                Q(customer__customer_code__icontains=q) |
                Q(customer__name__icontains=q)
            ).distinct()

    return render(request, 'contract_copyright_search.html', {'contracts': contracts, 'copyrights': copyrights, 'q': q})


def contract_business_search(request):
    q = request.GET.get('q', '').strip()
    contracts = Contract.objects.filter(service_type='dkkd')
    if q:
        business = BusinessRegistrationService.objects.filter(tax_code__icontains=q).first()
        if business:
            return redirect('business_detail', business_id=business.id)
        contracts = contracts.filter(
            Q(contract_no__icontains=q) |
            Q(customer__customer_code__icontains=q) |
            Q(customer__name__icontains=q)
        )
    return render(request, 'contract_business_search.html', {'contracts': contracts.order_by('-created_at'), 'q': q})


def contract_investment_search(request):
    q = request.GET.get('q', '').strip()
    contracts = Contract.objects.filter(service_type='dautu')
    if q:
        investment = InvestmentService.objects.filter(project_code__icontains=q).first()
        if investment:
            return redirect('investment_detail', investment_id=investment.id)
        contracts = contracts.filter(
            Q(contract_no__icontains=q) |
            Q(customer__customer_code__icontains=q) |
            Q(customer__name__icontains=q)
        )
    return render(request, 'contract_investment_search.html', {'contracts': contracts.order_by('-created_at'), 'q': q})


def contract_other_service_search(request):
    q = request.GET.get('q', '').strip()
    contracts = Contract.objects.filter(service_type='khac')
    if q:
        contracts = contracts.filter(
            Q(contract_no__icontains=q) |
            Q(customer__customer_code__icontains=q) |
            Q(customer__name__icontains=q)
        )
    return render(request, 'contract_other_service_search.html', {'contracts': contracts.order_by('-created_at'), 'q': q})


def trademark_search(request):
    q = request.GET.get('q', '').strip()
    if q:
        trademark = TrademarkService.objects.filter(app_no__icontains=q).first()
        if trademark:
            return redirect('trademark_detail', trademark_id=trademark.id)
        return render(request, 'trademark_search.html', {'q': q, 'not_found': True})
    return render(request, 'trademark_search.html', {'q': q})


def trademark_detail(request, trademark_id):
    trademark = get_object_or_404(TrademarkService, id=trademark_id)
    return render(request, 'trademark_detail.html', {'trademark': trademark, 'contract': trademark.contract})


def copyright_search(request):
    q = request.GET.get('q', '').strip()
    if q:
        copyright = CopyrightService.objects.filter(certificate_no__icontains=q).first()
        if copyright:
            return redirect('copyright_detail', copyright_id=copyright.id)
        return render(request, 'copyright_search.html', {'q': q, 'not_found': True})
    return render(request, 'copyright_search.html', {'q': q})


def copyright_detail(request, copyright_id):
    copyright = get_object_or_404(CopyrightService, id=copyright_id)
    return render(request, 'copyright_detail.html', {'copyright': copyright, 'contract': copyright.contract})


def business_search(request):
    q = request.GET.get('q', '').strip()
    if q:
        business = BusinessRegistrationService.objects.filter(tax_code__icontains=q).first()
        if business:
            return redirect('business_detail', business_id=business.id)
        return render(request, 'business_search.html', {'q': q, 'not_found': True})
    return render(request, 'business_search.html', {'q': q})


def business_detail(request, business_id):
    business = get_object_or_404(BusinessRegistrationService, id=business_id)
    return render(request, 'business_detail.html', {'business': business, 'contract': business.contract})


def investment_search(request):
    q = request.GET.get('q', '').strip()
    if q:
        investment = InvestmentService.objects.filter(project_code__icontains=q).first()
        if investment:
            return redirect('investment_detail', investment_id=investment.id)
        return render(request, 'investment_search.html', {'q': q, 'not_found': True})
    return render(request, 'investment_search.html', {'q': q})


def investment_detail(request, investment_id):
    investment = get_object_or_404(InvestmentService, id=investment_id)
    return render(request, 'investment_detail.html', {'investment': investment, 'contract': investment.contract})


# ===============================================
# FILE ĐÍNH KÈM
# ===============================================
def delete_certificate(request, pk):
    cert = get_object_or_404(Certificate, pk=pk)
    if cert.file:
        cert.file.delete(save=False)
    contract_id = cert.content_object.contract.id
    cert.delete()
    return redirect("contract_detail", contract_id)


def upload_certificate(request):
    if request.method == 'POST':
        service_type = request.POST.get('service_type')
        service_id   = request.POST.get('service_id')
        file         = request.FILES.get('file')
        name         = request.POST.get('name')

        if not file:
            messages.error(request, 'Chưa chọn file')
            return redirect(request.META.get('HTTP_REFERER'))

        MODEL_MAP = {
            'nhanhieu': TrademarkService,
            'banquyen': CopyrightService,
            'dkkd':     BusinessRegistrationService,
            'dautu':    InvestmentService,
            'khac':     OtherService,
        }

        model   = MODEL_MAP.get(service_type)
        service = get_object_or_404(model, id=service_id)

        Certificate.objects.create(content_object=service, name=name, file=file)

        messages.success(request, 'Đã thêm tài liệu')

        # 🔔 GỬI MAIL UPLOAD FILE
        _send_email(
            to_email=service.contract.customer.email,
            subject=f"[IPSHIELD] Tài liệu mới – {service.contract.contract_no}",
            message=f"""Kính gửi {service.contract.customer.name},

Một tài liệu mới vừa được thêm vào hợp đồng {service.contract.contract_no}.

Vui lòng đăng nhập portal để xem.

Trân trọng,
IPShield
""",
        )

        return redirect(request.META.get('HTTP_REFERER'))


# ===============================================
# HỒ SƠ NHÂN VIÊN
# ===============================================
from .models import UserProfile


@login_required
def profile_view(request):
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=profile, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Cập nhật hồ sơ thành công!')
            return redirect('profile')
        else:
            for errors in form.errors.values():
                for error in errors:
                    messages.error(request, error)
    else:
        form = UserProfileForm(instance=profile, user=request.user)
    return render(request, 'profile.html', {'form': form, 'profile': profile})


# ===============================================
# PORTAL KHÁCH HÀNG
# ===============================================
from .decorators import customer_login_required


def customer_login(request):
    if request.customer:
        return redirect('portal_dashboard')
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        customer_code = request.POST.get('customer_code', '').strip()
        password      = request.POST.get('password', '').strip()
        try:
            customer = Customer.objects.get(customer_code=customer_code)
            if customer.check_password(password):
                request.session['customer_id'] = customer.id
                customer.last_login = timezone.now()
                customer.save(update_fields=['last_login'])
                _log_activity(customer, 'login', request=request)
                return redirect('portal_dashboard')
            else:
                messages.error(request, '❌ Mật khẩu không đúng.')
        except Customer.DoesNotExist:
            messages.error(request, '❌ Mã khách hàng không tồn tại.')

    return render(request, 'customer_login.html')


def customer_logout(request):
    if 'customer_id' in request.session:
        if 'customer_id' in request.session:
            try:
                cust = Customer.objects.get(id=request.session['customer_id'])
                _log_activity(cust, 'logout', request=request)
            except Customer.DoesNotExist:
                pass
        del request.session['customer_id']
    return redirect('login')


@customer_login_required
def portal_dashboard(request):
    customer = request.customer
    contracts = Contract.objects.filter(customer=customer).order_by('-created_at')

    from collections import defaultdict
    grouped = defaultdict(list)

    for c in contracts:
        grouped[c.service_type].append(c)

    # ✅ thêm grouped_tabs cho template
    grouped_tabs = [('all', contracts)]

    for key in ['nhanhieu', 'banquyen', 'dkkd', 'dautu', 'khac']:
        if grouped.get(key):
            grouped_tabs.append((key, grouped[key]))

    return render(request, 'portal/dashboard.html', {
        'customer': customer,
        'contracts': contracts,
        'grouped': grouped,
        'grouped_tabs': grouped_tabs,   # ✅ QUAN TRỌNG
        'total': contracts.count(),
        'completed': contracts.filter(status='completed').count(),
        'pending': contracts.exclude(status='completed').count(),
    })


@customer_login_required
def portal_contract_detail(request, contract_id):
    customer  = request.customer
    contract  = get_object_or_404(Contract, id=contract_id, customer=customer)
    installments = contract.installments.all()

    service = None
    if contract.service_type == 'nhanhieu':
        service = contract.trademarks.all()
    elif contract.service_type == 'banquyen':
        service = contract.copyrights.all()
    elif contract.service_type == 'dkkd':
        service = BusinessRegistrationService.objects.filter(contract=contract)
    elif contract.service_type == 'dautu':
        service = InvestmentService.objects.filter(contract=contract)
    else:
        service = OtherService.objects.filter(contract=contract)
    _log_activity(customer, 'view_contract', note=contract.contract_no, request=request)
    return render(request, 'portal/contract_detail.html', {
        'contract': contract, 'service': service,
        'installments': installments, 'customer': customer,
    })


@customer_login_required
def portal_customer_profile(request):
    customer = request.customer

    if request.method == 'POST':

        # 🔔 XỬ LÝ YÊU CẦU CHỈNH SỬA THÔNG TIN
        if request.POST.get('request_type') == 'edit_profile':
            content = request.POST.get('request_content', '').strip()
            if content:
                _log_activity(
                    customer,
                    'support_request',
                    note=f"[Yêu cầu chỉnh sửa thông tin] {content}",
                    request=request
                )
                _send_email(
                    to_email=settings.DEFAULT_FROM_EMAIL,
                    subject=f"[IPSHIELD] Yêu cầu chỉnh sửa thông tin – {customer.customer_code}",
                    message=f"""Khách hàng {customer.name} ({customer.customer_code}) vừa gửi yêu cầu chỉnh sửa thông tin:

  • Email  : {customer.email}
  • SĐT    : {customer.phone}

Nội dung yêu cầu:
{content}
""",
                )
                messages.success(request, '✅ Yêu cầu chỉnh sửa đã được ghi nhận! Chúng tôi sẽ liên hệ sớm nhất.')
            else:
                messages.error(request, '❌ Vui lòng nhập nội dung yêu cầu.')
            return redirect('portal_customer_profile')

        # ĐỔI MẬT KHẨU / CẬP NHẬT HỒ SƠ
        new_password     = request.POST.get('new_password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()

        if request.FILES.get('avatar'):
            customer.avatar = request.FILES['avatar']

        if new_password:
            if new_password != confirm_password:
                messages.error(request, '❌ Mật khẩu xác nhận không khớp!')
                return redirect('portal_customer_profile')
            customer.set_password(new_password)

        customer.save()
        action_type = 'change_password' if new_password else 'update_profile'
        _log_activity(customer, action_type, request=request)

        _send_email(
            to_email=customer.email,
            subject="[IPSHIELD] Hồ sơ của bạn đã được cập nhật",
            message=f"""Kính gửi {customer.name},

Thông tin hồ sơ cá nhân của bạn đã được cập nhật lúc {timezone.now().strftime("%d/%m/%Y %H:%M")}.

Nếu bạn không thực hiện thay đổi này, vui lòng liên hệ ngay với chúng tôi.

Trân trọng,
IPShield
""",
        )

        messages.success(request, '✅ Cập nhật hồ sơ thành công!')
        return redirect('portal_customer_profile')

    return render(request, 'portal/customer_profile.html', {'customer': customer})




@login_required
def dashboard(request):
    import datetime
    from django.db.models import Count
    from django.db.models.functions import TruncDate

    today       = datetime.date.today()
    year_range  = range(today.year - 3, today.year + 1)
    month_range = range(1, 13)

    action_filter = request.GET.get('action', '')
    q             = request.GET.get('q', '').strip()
    date_from     = request.GET.get('date_from', '')
    date_to       = request.GET.get('date_to', '')

    logs = CustomerActivityLog.objects.select_related('customer').order_by('-created_at')

    if action_filter:
        logs = logs.filter(action=action_filter)
    if q:
        logs = logs.filter(
            Q(customer__customer_code__icontains=q) |
            Q(customer__name__icontains=q)
        )
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)

    # ── Xuất theo ngày: toàn bộ hoạt động của mọi KH trong ngày được chọn ──
    if request.GET.get('export') == 'excel_by_date':
        export_date = request.GET.get('export_date', str(today))
        export_logs = CustomerActivityLog.objects.select_related('customer').filter(
            created_at__date=export_date
        ).order_by('customer__customer_code', 'created_at')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hoạt động theo ngày"
        ws.append(['STT', 'Mã KH', 'Tên KH', 'Hành động', 'Ghi chú', 'IP', 'Thời gian'])
        for idx, log in enumerate(export_logs, 1):
            ws.append([
                idx,
                log.customer.customer_code,
                log.customer.name,
                log.get_action_display(),
                log.note or '',
                log.ip_address or '',
                log.created_at.strftime('%H:%M:%S — %d/%m/%Y'),
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="hoat_dong_ngay_{export_date}.xlsx"'
        )
        wb.save(response)
        return response

    # ── Xuất theo KH: hoạt động của một KH cụ thể trong ngày được chọn ──
    if request.GET.get('export') == 'excel_by_customer':
        export_date        = request.GET.get('export_date', str(today))
        export_customer_id = request.GET.get('export_customer_id', '').strip()

        export_logs = CustomerActivityLog.objects.select_related('customer').filter(
            created_at__date=export_date
        ).order_by('created_at')

        if export_customer_id:
            export_logs = export_logs.filter(customer__id=export_customer_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Theo KH"
        ws.append(['STT', 'Mã KH', 'Tên KH', 'Hành động', 'Ghi chú', 'IP', 'Thời gian'])
        for idx, log in enumerate(export_logs, 1):
            ws.append([
                idx,
                log.customer.customer_code,
                log.customer.name,
                log.get_action_display(),
                log.note or '',
                log.ip_address or '',
                log.created_at.strftime('%H:%M:%S — %d/%m/%Y'),
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="hoat_dong_kh_{export_customer_id}_{export_date}.xlsx"'
        )
        wb.save(response)
        return response

    # ── Thống kê nhanh ──
    stats = CustomerActivityLog.objects.values('action').annotate(total=Count('id'))
    stats_dict = {s['action']: s['total'] for s in stats}

    # Danh sách KH để chọn khi xuất theo KH
    customers = CustomerActivityLog.objects.select_related('customer') \
        .values('customer__id', 'customer__customer_code', 'customer__name') \
        .distinct().order_by('customer__customer_code')

    return render(request, 'dashboard.html', {
        'logs':           logs[:200],
        'action_filter':  action_filter,
        'q':              q,
        'date_from':      date_from,
        'date_to':        date_to,
        'stats':          stats_dict,
        'action_choices': CustomerActivityLog.ACTION_CHOICES,
        'today':          str(today),
        'current_month':  today.month,
        'current_year':   today.year,
        'year_range':     year_range,
        'month_range':    month_range,
        'customers':      customers,   # <-- thêm mới
    })


# LOG ACTIVITIES
@login_required
def dashboard_stats_api(request):
    import datetime
    from django.db.models import Count

    period = request.GET.get('period', 'month')
    logs   = CustomerActivityLog.objects.all()

    if period == 'day':
        date = request.GET.get('date', str(datetime.date.today()))
        logs = logs.filter(created_at__date=date)
    elif period == 'month':
        month = int(request.GET.get('month', datetime.date.today().month))
        year  = int(request.GET.get('year',  datetime.date.today().year))
        logs  = logs.filter(created_at__month=month, created_at__year=year)
    elif period == 'year':
        year = int(request.GET.get('year', datetime.date.today().year))
        logs = logs.filter(created_at__year=year)

    stats = logs.values('action').annotate(total=Count('id'))
    return JsonResponse({s['action']: s['total'] for s in stats})

# API KHÁCH HÀNG
@login_required
def dashboard_customer_stats_api(request):
    from django.db.models import Count
    import datetime
    period = request.GET.get('period', 'month')
    customers = Customer.objects.all()

    today = datetime.date.today()

    # ─────────────────────────────
    # 📅 FILTER THEO THỜI GIAN
    # ─────────────────────────────
    if period == 'day':
        date_str = request.GET.get('date')
        if date_str:
            customers = customers.filter(register_year=date_str)

    elif period == 'month':
        month = request.GET.get('month')
        year  = request.GET.get('year')

        if month and year:
            customers = customers.filter(
                register_year__month=int(month),
                register_year__year=int(year)
            )

    elif period == 'year':
        year = request.GET.get('year')
        if year:
            customers = customers.filter(register_year__year=int(year))

    # ─────────────────────────────
    # 📊 THỐNG KÊ
    # ─────────────────────────────

    # Tổng KH
    total = customers.count()

    # KH mới (từ năm 2026)
    new_customers = customers.filter(register_year__year=today.year).count()

    # group theo status
    status_stats = Customer.objects.values('status').annotate(total=Count('id'))
    status_data = {s['status']: s['total'] for s in status_stats}

    # ─────────────────────────────
    # 📦 RESPONSE
    # ─────────────────────────────
    data = {
        "total": total,
        "new": new_customers,
        "approved": status_data.get('approved', 0),     # chờ duyệt
        "pending": status_data.get('pending', 0),       # đang duyệt
        "completed": status_data.get('completed', 0),   # hoàn tất
    }
    return JsonResponse(data)

@login_required
def dashboard_trademark_stats_api(request):
    import datetime
    from django.http import JsonResponse
    from .models import TrademarkService

    period = request.GET.get('period', 'month')

    qs = TrademarkService.objects.all()

    date_str = request.GET.get('date')
    month = request.GET.get('month')
    year = request.GET.get('year')

    def filter_by_date(queryset, field):
        if period == 'day' and date_str:
            return queryset.filter(**{field: date_str})

        elif period == 'month' and month and year:
            return queryset.filter(
                **{
                    f"{field}__month": int(month),
                    f"{field}__year": int(year)
                }
            )

        elif period == 'year' and year:
            return queryset.filter(**{f"{field}__year": int(year)})

        return queryset

    # ✅ TOTAL (nên thống nhất theo filing_date)
    total = filter_by_date(qs, 'filing_date').count()

    # ✅ ĐÃ NỘP ĐƠN
    filing_date_count = filter_by_date(qs, 'filing_date') \
        .filter(filing_date__isnull=False).count()

    # ✅ ĐÃ CẤP
    decision_date_count = filter_by_date(qs, 'decision_date') \
        .filter(decision_date__isnull=False).count()

    # ✅ HỢP LỆ
    valid_date_count = filter_by_date(qs, 'valid_date') \
        .filter(valid_date__isnull=False).count()

    # ✅ BỊ TỪ CHỐI (KHÔNG filter theo date)
    deny_document_count = qs.filter(deny_document=True).count()

    return JsonResponse({
        'total': total,
        "filled": filing_date_count,
        "decision": decision_date_count,
        "valid": valid_date_count,
        "deny": deny_document_count,
    })

    
# YÊU CẦU HỖ TRỢ
@customer_login_required
def portal_support_request(request, contract_id):
    if request.method != 'POST':
        return redirect('portal_contract_detail', contract_id=contract_id)

    customer       = request.customer
    contract       = get_object_or_404(Contract, id=contract_id, customer=customer)
    subject        = request.POST.get('subject', '').strip()
    message        = request.POST.get('message', '').strip()
    contact_method = request.POST.get('contact_method', 'email')

    if not subject or not message:
        return JsonResponse({'ok': False, 'error': 'Vui lòng nhập đầy đủ thông tin'}, status=400)

    _log_activity(
        customer,
        'support_request',
        note=f"[{contract.contract_no}] {subject} | {message} | Liên hệ: {contact_method}",
        request=request
    )
    _send_email(
                    to_email=settings.DEFAULT_FROM_EMAIL,
                    subject=f"[IPSHIELD] Yêu cầu chỉnh sửa thông tin – {customer.customer_code}",
                    message=f"""Khách hàng {customer.name} ({customer.customer_code}) vừa gửi yêu cầu chỉnh sửa thông tin:

  • Email  : {customer.email}
  • SĐT    : {customer.phone}

Tiêu đề: {subject}
Nội dung yêu cầu:
{message}
Liên hệ qua: {contact_method}

Trân trọng, 
{ customer.name }
""",
                )
    messages.success(request, '✅ Yêu cầu chỉnh sửa đã được ghi nhận! Chúng tôi sẽ liên hệ sớm nhất.')


    return JsonResponse({'ok': True})

